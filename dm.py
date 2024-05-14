import openpyxl

# 텍스트 파일에서 데이터 읽기
with open('data.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

# 엑셀 파일 생성
wb = openpyxl.Workbook()
sheet = wb.active

row = 1
for i in range(0, len(lines), 2):  # 파일 이름과 파일 크기가 번갈아 나오므로 2씩 증가
    filename = lines[i].strip()  # 파일 이름은 현재 줄
    if i + 1 < len(lines):  # 파일 크기를 읽어올 줄이 리스트 범위 내에 있는지 확인
        size = lines[i + 1].strip()  # 파일 크기는 다음 줄
    else:
        size = ""  # 파일 크기가 없는 경우

    # 엑셀에 데이터 쓰기
    sheet.cell(row=row, column=1, value=filename)
    sheet.cell(row=row, column=2, value=size)

    row += 1

# 엑셀 파일 저장
wb.save('data.xlsx')
